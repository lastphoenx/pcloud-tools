#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
analyze_manifest_duplicates.py - Analysiere JSON-Manifest auf SHA256-Duplikate

Beispielaufruf:
  python analyze_manifest_duplicates.py \
    --manifest /srv/pcloud-archive/manifests/2026-04-10-075334.json \
    --output duplicates.xlsx
"""

import json
import argparse
import sys
from pathlib import Path
from typing import List, Dict, Any
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def load_manifest(manifest_path: str) -> Dict[str, Any]:
    """Lade JSON-Manifest"""
    with open(manifest_path, 'r', encoding='utf-8') as f:
        return json.load(f)


def extract_file_data(manifest: Dict[str, Any]) -> pd.DataFrame:
    """
    Extrahiere Datei-Informationen aus dem Manifest
    
    Returns:
        DataFrame mit Spalten: full_path, directory, filename, sha256, size, mtime
    """
    items = manifest.get('items', [])
    
    # Nur Dateien (nicht dirs oder symlinks)
    files = [item for item in items if item.get('type') == 'file']
    
    records = []
    for item in files:
        relpath = item.get('relpath', '')
        sha256 = item.get('sha256', '')
        
        # Nur Dateien mit SHA256 aufnehmen
        if not sha256:
            continue
        
        # Pfad aufteilen
        path_obj = Path(relpath)
        directory = str(path_obj.parent) if path_obj.parent != Path('.') else ''
        filename = path_obj.name
        
        records.append({
            'full_path': relpath,
            'directory': directory,
            'filename': filename,
            'sha256': sha256,
            'size': item.get('size', 0),
            'mtime': item.get('mtime', 0)
        })
    
    return pd.DataFrame(records)


def analyze_duplicates(df: pd.DataFrame) -> pd.DataFrame:
    """
    Füge Duplikat-Informationen und Duplikat-ID hinzu
    
    Zusätzliche Spalten:
    - duplicate_id: Fortlaufende ID für jede Duplikat-Gruppe
    - duplicate_count: Anzahl Dateien mit gleichem SHA256
    - is_duplicate: Ob Datei ein Duplikat ist (count > 1)
    - wasted_space: Verschwendeter Platz (size * (count - 1))
    """
    # Zähle Vorkommen pro SHA256
    sha256_counts = df['sha256'].value_counts().to_dict()
    df['duplicate_count'] = df['sha256'].map(sha256_counts)
    df['is_duplicate'] = df['duplicate_count'] > 1
    
    # Duplikat-ID: fortlaufende Nummer für jede Duplikat-Gruppe
    duplicate_sha256s = df[df['is_duplicate']]['sha256'].unique()
    sha256_to_id = {sha: idx + 1 for idx, sha in enumerate(sorted(duplicate_sha256s))}
    df['duplicate_id'] = df['sha256'].map(sha256_to_id)
    df['duplicate_id'] = df['duplicate_id'].fillna(0).astype(int)
    
    # Platzverschwendung berechnen (KORREKT pro Zeile, nicht mehrfach gezählt):
    # Strategie: Sortiere pro SHA256-Gruppe, erste Instanz = 0 (würden wir behalten),
    # alle weiteren = size (sind Verschwendung)
    df['wasted_space'] = 0
    
    # Für jede SHA256-Gruppe: Alle außer der ersten sind Verschwendung
    for sha256, group in df[df['is_duplicate']].groupby('sha256'):
        # Erste Instanz behalten (wasted_space = 0)
        # Weitere Instanzen: jeweils 'size' ist verschwendet
        indices = group.index.tolist()
        if len(indices) > 1:
            # Indices ab Position 1 (2., 3., ... Kopie) sind Verschwendung
            df.loc[indices[1:], 'wasted_space'] = df.loc[indices[1:], 'size']
    
    return df


def format_bytes(bytes_val: int) -> str:
    """Formatiere Bytes als lesbare Größe"""
    for unit in ['B', 'KB', 'MB', 'GB', 'TB']:
        if bytes_val < 1024.0 or unit == 'TB':
            return f"{bytes_val:.1f} {unit}"
        bytes_val /= 1024.0


def create_space_analysis(df: pd.DataFrame) -> pd.DataFrame:
    """
    Erstelle Platzersparnis-Analyse
    
    Returns:
        DataFrame mit Top-Duplikaten sortiert nach Platzverschwendung
    """
    # Gruppiere nach SHA256 (nur Duplikate)
    df_dupes = df[df['is_duplicate']].copy()
    
    if df_dupes.empty:
        return pd.DataFrame()
    
    # Pro Duplikat-Gruppe: Summe des wasted_space (= size × (count - 1))
    analysis = df_dupes.groupby('sha256').agg({
        'duplicate_id': 'first',
        'filename': 'first',  # Beispiel-Dateiname
        'size': 'first',  # Einzeldateigröße
        'duplicate_count': 'first',
        'wasted_space': 'sum'  # SUMME des wasted_space (korrekt, da pro Zeile nur eigener Anteil)
    }).reset_index()
    
    # Readable Formate hinzufügen
    analysis['size_readable'] = analysis['size'].apply(format_bytes)
    analysis['wasted_space_readable'] = analysis['wasted_space'].apply(format_bytes)
    
    # Sortiere nach Platzverschwendung
    analysis = analysis.sort_values('wasted_space', ascending=False)
    
    # Spalten umbenennen
    analysis = analysis.rename(columns={
        'duplicate_id': 'Duplikat-ID',
        'filename': 'Beispiel-Datei',
        'size': 'Einzelgröße (Bytes)',
        'size_readable': 'Einzelgröße',
        'duplicate_count': 'Anzahl Kopien',
        'wasted_space': 'Verschwendet (Bytes)',
        'wasted_space_readable': 'Verschwendet',
        'sha256': 'SHA256'
    })
    
    return analysis[['Duplikat-ID', 'Beispiel-Datei', 'Einzelgröße', 'Anzahl Kopien', 'Verschwendet', 'SHA256']]


def apply_excel_formatting(excel_path: str, stats: Dict[str, Any]) -> None:
    """
    Wende Formatierung auf Excel-Datei an
    - Farbige Zeilen für Duplikate
    - Auto-Spaltenbreite
    - Header-Formatierung
    - Zahlenformate
    """
    wb = load_workbook(excel_path)
    
    # Farben definieren
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    duplicate_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
    
    border_thin = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )
    
    # Sheet 1: Alle Dateien
    if 'Alle Dateien' in wb.sheetnames:
        ws = wb['Alle Dateien']
        
        # Header formatieren
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border_thin
        
        # Auto-Spaltenbreite
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 60)  # Max 60 Zeichen
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Duplikate farbig markieren (is_duplicate = True)
        is_duplicate_col = None
        for idx, cell in enumerate(ws[1], 1):
            if cell.value == 'is_duplicate':
                is_duplicate_col = idx
                break
        
        if is_duplicate_col:
            for row_idx, row in enumerate(ws.iter_rows(min_row=2), 2):
                if row[is_duplicate_col - 1].value is True:
                    for cell in row:
                        if not cell.fill.start_color.rgb or cell.fill.start_color.rgb == '00000000':
                            cell.fill = duplicate_fill
        
        # Freeze Panes (erste Zeile)
        ws.freeze_panes = 'A2'
    
    # Sheet 2: Nur Duplikate (gleiche Formatierung)
    if 'Nur Duplikate' in wb.sheetnames:
        ws = wb['Nur Duplikate']
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border_thin
        
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 60)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Alle Zeilen farbig (sind ja alles Duplikate)
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if not cell.fill.start_color.rgb or cell.fill.start_color.rgb == '00000000':
                    cell.fill = duplicate_fill
        
        ws.freeze_panes = 'A2'
    
    # Sheet 3: Platzersparnis-Analyse
    if 'Platzersparnis-Analyse' in wb.sheetnames:
        ws = wb['Platzersparnis-Analyse']
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border_thin
        
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 60)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        ws.freeze_panes = 'A2'
    
    # Sheet 4: Statistik
    if 'Statistik' in wb.sheetnames:
        ws = wb['Statistik']
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border_thin
        
        # Spaltenbreiten anpassen
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 25
        
        # Werte rechts ausrichten
        for row in ws.iter_rows(min_row=2):
            if len(row) > 1:
                row[1].alignment = Alignment(horizontal='right')
    
    wb.save(excel_path)


def export_to_excel(df: pd.DataFrame, output_path: str, snapshot_name: str) -> None:
    """
    Exportiere nach Excel mit Formatierung
    """
    # Sortiere nach SHA256 und dann nach Pfad
    df_sorted = df.sort_values(['duplicate_id', 'sha256', 'full_path'])
    
    # Spalten-Reihenfolge festlegen
    column_order = [
        'duplicate_id', 'sha256', 'full_path', 'directory', 'filename',
        'size', 'duplicate_count', 'is_duplicate', 'wasted_space', 'mtime'
    ]
    df_sorted = df_sorted[column_order]
    
    # Statistiken berechnen
    total_files = len(df)
    total_duplicates = len(df[df['is_duplicate']])
    unique_hashes = df['sha256'].nunique()
    duplicate_groups = df[df['is_duplicate']]['sha256'].nunique()
    total_wasted = df['wasted_space'].sum()
    
    # Erstelle Excel-Writer
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # Sheet 1: Alle Dateien
        df_sorted.to_excel(writer, sheet_name='Alle Dateien', index=False)
        
        # Sheet 2: Nur Duplikate
        df_dupes = df_sorted[df_sorted['is_duplicate']]
        if not df_dupes.empty:
            df_dupes.to_excel(writer, sheet_name='Nur Duplikate', index=False)
        
        # Sheet 3: Platzersparnis-Analyse
        space_analysis = create_space_analysis(df)
        if not space_analysis.empty:
            space_analysis.to_excel(writer, sheet_name='Platzersparnis-Analyse', index=False)
        
        # Sheet 4: Statistik
        stats = {
            'Metrik': [
                'Snapshot',
                'Gesamt Dateien',
                'Dateien mit SHA256',
                'Anzahl Duplikate',
                'Unique SHA256-Hashes',
                'Duplikate-Gruppen',
                'Verschwendeter Platz (Bytes)',
                'Verschwendeter Platz'
            ],
            'Wert': [
                snapshot_name,
                total_files,
                len(df[df['sha256'] != '']),
                total_duplicates,
                unique_hashes,
                duplicate_groups,
                total_wasted,
                format_bytes(total_wasted)
            ]
        }
        pd.DataFrame(stats).to_excel(writer, sheet_name='Statistik', index=False)
    
    # Formatierung anwenden
    apply_excel_formatting(output_path, {
        'total_wasted': total_wasted,
        'duplicate_groups': duplicate_groups
    })
    
    print(f"✓ Excel exportiert nach: {output_path}")


def main():
    parser = argparse.ArgumentParser(
        description='Analysiere JSON-Manifest auf SHA256-Duplikate'
    )
    parser.add_argument(
        '--manifest',
        required=True,
        help='Pfad zum JSON-Manifest (z.B. /srv/pcloud-archive/manifests/2026-04-10-075334.json)'
    )
    parser.add_argument(
        '--output',
        required=True,
        help='Ausgabe-Datei (Excel, z.B. duplicates.xlsx)'
    )
    
    args = parser.parse_args()
    
    # Validierung
    if not Path(args.manifest).exists():
        print(f"✗ Fehler: Manifest nicht gefunden: {args.manifest}", file=sys.stderr)
        sys.exit(1)
    
    # Verarbeitung
    print(f"Lade Manifest: {args.manifest}")
    manifest = load_manifest(args.manifest)
    
    snapshot_name = manifest.get('snapshot', '?')
    version = manifest.get('version', '?')
    
    print(f"Snapshot: {snapshot_name}")
    print(f"Version: {version}")
    
    print("Extrahiere Datei-Daten...")
    df = extract_file_data(manifest)
    print(f"  → {len(df)} Dateien mit SHA256 gefunden")
    
    print("Analysiere Duplikate...")
    df = analyze_duplicates(df)
    dupes_count = len(df[df['is_duplicate']])
    dupes_groups = df[df['is_duplicate']]['sha256'].nunique()
    total_wasted = df['wasted_space'].sum()
    print(f"  → {dupes_count} Duplikate in {dupes_groups} Gruppen gefunden")
    print(f"  → {format_bytes(total_wasted)} Platz verschwendet")
    
    print(f"Exportiere nach Excel...")
    export_to_excel(df, args.output, snapshot_name)
    
    print("\n" + "="*60)
    print("ZUSAMMENFASSUNG")
    print("="*60)
    print(f"Gesamt Dateien:       {len(df)}")
    print(f"Duplikate:            {dupes_count}")
    print(f"Duplikat-Gruppen:     {dupes_groups}")
    print(f"Unique Hashes:        {df['sha256'].nunique()}")
    print(f"Verschwendeter Platz: {format_bytes(total_wasted)}")
    print("="*60)
    
    # Top 5 Platzfresser
    if dupes_count > 0:
        print("\nTop 5 Platzfresser:")
        top5 = df[df['is_duplicate']].nlargest(5, 'wasted_space')
        for idx, row in top5.iterrows():
            print(f"  {format_bytes(row['wasted_space']):>12} | "
                  f"{row['duplicate_count']}× | {row['filename']}")
    
    print(f"\n✓ Fertig! Öffne {args.output}")


if __name__ == '__main__':
    main()
